# -*- coding: utf-8 -*-
"""Полный анализ структуры ДДС.xlsx без изменений файла."""
import openpyxl
import json
from collections import defaultdict

PATH = r"C:\finance\data\ДДС.xlsx"
wb = openpyxl.load_workbook(PATH, data_only=True)

def nonempty_row(ws, r, max_col=None):
    mc = max_col or ws.max_column
    return any(ws.cell(r, c).value is not None for c in range(1, mc + 1))

def scan_sheet(name):
    ws = wb[name]
    rows_with_a = []
    for r in range(1, min(ws.max_row, 200) + 1):
        a = ws.cell(r, 1).value
        if a and str(a).strip():
            rows_with_a.append((r, str(a).strip()[:70]))
    return {
        "dims": f"{ws.max_column}x{ws.max_row}",
        "nonempty": sum(1 for r in range(1, ws.max_row + 1) if nonempty_row(ws, r)),
        "labels_col_a": rows_with_a[:80],
    }

# Snapshot columns in main sheet
ws = wb["Основные данные за месяц"]
snapshots = []
for c in range(1, ws.max_column + 1, 2):
    lbl = ws.cell(3, c).value
    val = ws.cell(3, c + 1).value if c + 1 <= ws.max_column else None
    if lbl and "Остаток" in str(lbl):
        snapshots.append({"col": openpyxl.utils.get_column_letter(c + 1), "label": lbl, "value": val})

# P&L block AC
pnl = []
for r in range(3, 25):
    ac = ws.cell(r, 29).value  # AC
    ae = ws.cell(r, 31).value  # AE
    if ac:
        pnl.append({"row": r, "label": ac, "value": ae})

# Balance AG-AJ
balance = []
for r in range(3, 15):
    ag = ws.cell(r, 33).value
    ah = ws.cell(r, 34).value
    ai = ws.cell(r, 35).value
    aj = ws.cell(r, 36).value
    if ag or ai:
        balance.append({"row": r, "ag": ag, "ah": ah, "ai": ai, "aj": aj})

# BDDS months row 8
ws_b = wb["БДДС"]
months = []
for c in range(1, ws_b.max_column + 1):
    v = ws_b.cell(8, c).value
    if v and "2026" in str(v):
        months.append({"col": openpyxl.utils.get_column_letter(c), "month": str(v).strip()})

# CF weeks
ws_cf = wb["CF по неделям"]
weeks_in = [ws_cf.cell(21, c).value for c in range(2, 18) if ws_cf.cell(21, c).value]
cf_sections = []
for r in range(1, 140):
    a = ws_cf.cell(r, 1).value
    if a and (str(a)[0].isdigit() or "Cash" in str(a) or "Статьи" in str(a) or str(a) == "Итог"):
        cf_sections.append((r, str(a)[:60]))

out = {
    "sheets": {n: scan_sheet(n) for n in wb.sheetnames},
    "snapshots": snapshots,
    "pnl_blocks": pnl,
    "mgmt_balance": balance,
    "bdds_months": months,
    "cf_weeks_sample": weeks_in[:14],
    "cf_sections": cf_sections,
}
print(json.dumps(out, ensure_ascii=False, indent=2))
